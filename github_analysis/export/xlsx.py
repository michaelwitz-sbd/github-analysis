from __future__ import annotations

from pathlib import Path
from typing import Iterable


def read_tsv_table(path: Path) -> tuple[list[str], list[list[str]]]:
    """Read a report TSV, skipping preamble/footer note lines."""
    lines = path.read_text(encoding="utf-8").splitlines()
    header: list[str] | None = None
    rows: list[list[str]] = []

    for line in lines:
        if not line.strip():
            continue
        if line.startswith("END NOTES"):
            break
        if "\t" not in line:
            continue
        cells = line.split("\t")
        if header is None:
            header = cells
            continue
        if len(cells) == len(header):
            rows.append(cells)

    if header is None:
        raise ValueError(f"no tabular header found in {path}")
    return header, rows


def _autosize_columns(ws, headers: list[str], rows: Iterable[list[str]]) -> None:
    from openpyxl.utils import get_column_letter

    widths = [len(header) for header in headers]
    for row in rows:
        for idx, value in enumerate(row):
            if idx < len(widths):
                widths[idx] = max(widths[idx], len(value))
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 60)


def export_workbook(
    *,
    summary_path: Path,
    detail_path: Path | None,
    output_path: Path,
    include_detail: bool,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for Excel export. Run: uv sync"
        ) from exc

    summary_headers, summary_rows = read_tsv_table(summary_path)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Individual Production"
    summary_ws.append(summary_headers)
    for row in summary_rows:
        summary_ws.append(row)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    _autosize_columns(summary_ws, summary_headers, summary_rows)

    if include_detail and detail_path is not None:
        detail_headers, detail_rows = read_tsv_table(detail_path)
        detail_ws = workbook.create_sheet("PR Detail")
        detail_ws.append(detail_headers)
        for row in detail_rows:
            detail_ws.append(row)
        for cell in detail_ws[1]:
            cell.font = Font(bold=True)
        _autosize_columns(detail_ws, detail_headers, detail_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
