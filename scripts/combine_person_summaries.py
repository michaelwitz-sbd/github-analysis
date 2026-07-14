#!/usr/bin/env python3
"""Combine per-repo *_person_summary.tsv files into one Excel workbook."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

# Allow running from repo root without install
_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from github_analysis.export.xlsx import read_tsv_table, _autosize_columns  # noqa: E402

COUNT_COLUMNS = {
    "prs_merged",
    "prs_reviewed",
    "prs_approved",
    "prs_authored",
    "prs_open",
    "prs_closed_unmerged",
}
WEIGHT_AUTHORED = {"avg_files_added_per_pr", "avg_files_changed_per_pr"}
WEIGHT_MERGED = {
    "min_hours_pr_created_to_merged",
    "max_hours_pr_created_to_merged",
    "avg_hours_pr_created_to_merged",
}

DEFAULT_TEAMS_CONFIG = _REPO_ROOT / "config" / "teams.yaml"


@dataclass(frozen=True)
class TeamMember:
    name: str
    user: str


@dataclass(frozen=True)
class Team:
    name: str
    members: list[TeamMember]
    team_banner: str = ""


def _parse_float(s: str) -> float | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(s: str) -> int:
    s = (s or "").strip()
    if not s:
        return 0
    return int(s)


_DATE_WINDOW_SUFFIX = re.compile(r"[-_]\d{4}-\d{2}-\d{2}_to_\d{4}-\d{2}-\d{2}$")


def _repo_label_from_summary_path(path: Path) -> str:
    stem = path.name
    if stem.endswith("_person_summary.tsv"):
        label = stem[: -len("_person_summary.tsv")]
    else:
        label = path.stem
    label = _DATE_WINDOW_SUFFIX.sub("", label)
    return label


def _sheet_title(name: str) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]"""
    title = re.sub(r"[:\\/?*\[\]]", "-", name)[:31]
    return title or "repo"


def load_teams_config(path: Path | None) -> list[Team]:
    """Load teams from YAML or JSON. Missing/empty path → no team sheets."""
    if path is None:
        return []
    path = path.expanduser()
    if not path.is_file():
        raise SystemExit(f"teams config not found: {path}")

    text = path.read_text(encoding="utf-8")
    suffix = path.suffix.lower()
    data: Any
    if suffix in {".yaml", ".yml"}:
        try:
            import yaml
        except ImportError as exc:
            raise SystemExit(
                "PyYAML required for .yaml teams config: uv sync --group excel"
            ) from exc
        data = yaml.safe_load(text)
    elif suffix == ".json":
        data = json.loads(text)
    else:
        raise SystemExit(f"teams config must be .yaml, .yml, or .json (got {path})")

    if not data:
        return []
    if not isinstance(data, dict) or "teams" not in data:
        raise SystemExit(f"teams config must be an object with a 'teams' list: {path}")

    teams_raw = data.get("teams") or []
    if not isinstance(teams_raw, list):
        raise SystemExit(f"'teams' must be a list in {path}")

    teams: list[Team] = []
    for i, raw in enumerate(teams_raw):
        if not isinstance(raw, dict):
            raise SystemExit(f"teams[{i}] must be an object in {path}")
        name = str(raw.get("name") or "").strip()
        if not name:
            raise SystemExit(f"teams[{i}] missing name in {path}")
        # Prefer team_banner; accept legacy team_header
        banner_raw = raw.get("team_banner")
        if banner_raw is None or str(banner_raw).strip() == "":
            banner_raw = raw.get("team_header")
        team_banner = str(banner_raw or name).strip() or name
        members_raw = raw.get("members") or []
        if not isinstance(members_raw, list) or not members_raw:
            raise SystemExit(f"teams[{i}] ({name}) must have a non-empty members list")
        members: list[TeamMember] = []
        for j, m in enumerate(members_raw):
            if not isinstance(m, dict):
                raise SystemExit(f"teams[{i}].members[{j}] must be an object")
            staff = str(m.get("name") or "").strip()
            user = str(m.get("user") or "").strip()
            if not staff or not user:
                raise SystemExit(
                    f"teams[{i}].members[{j}] needs both 'name' and 'user' in {path}"
                )
            members.append(TeamMember(name=staff, user=user))
        teams.append(Team(name=name, members=members, team_banner=team_banner))
    return teams


def _aggregate_totals(
    headers: list[str],
    repo_rows: list[tuple[str, list[list[str]]]],
) -> tuple[list[str], list[list[str]]]:
    """Roll up person-summary rows across repos — one row per user, same columns as source."""
    idx = {h: i for i, h in enumerate(headers)}
    by_user: dict[str, dict] = defaultdict(
        lambda: {
            "counts": defaultdict(int),
            "file_w": defaultdict(float),
            "file_n": defaultdict(float),
            "merge_w": defaultdict(float),
            "merge_n": defaultdict(float),
            "min_hours": {},
            "max_hours": {},
        }
    )

    for _repo, rows in repo_rows:
        for row in rows:
            user = row[idx["user"]]
            rec = by_user[user]
            for col in COUNT_COLUMNS:
                rec["counts"][col] += _parse_int(row[idx[col]])
            authored = _parse_int(row[idx["prs_authored"]])
            merged = _parse_int(row[idx["prs_merged"]])
            for col in WEIGHT_AUTHORED:
                val = _parse_float(row[idx[col]])
                if val is not None and authored > 0:
                    rec["file_w"][col] += val * authored
                    rec["file_n"][col] += authored
            for col in WEIGHT_MERGED:
                val = _parse_float(row[idx[col]])
                if col == "min_hours_pr_created_to_merged" and val is not None:
                    prev = rec["min_hours"].get(col)
                    rec["min_hours"][col] = val if prev is None else min(prev, val)
                elif col == "max_hours_pr_created_to_merged" and val is not None:
                    prev = rec["max_hours"].get(col)
                    rec["max_hours"][col] = val if prev is None else max(prev, val)
                elif col == "avg_hours_pr_created_to_merged" and val is not None and merged > 0:
                    rec["merge_w"][col] += val * merged
                    rec["merge_n"][col] += merged

    def _fmt_avg(w: float, n: float) -> str:
        if n <= 0:
            return ""
        return f"{w / n:.2f}"

    out: list[list[str]] = []
    for user in sorted(by_user.keys(), key=str.lower):
        rec = by_user[user]
        line = [""] * len(headers)
        line[idx["user"]] = user
        for col in COUNT_COLUMNS:
            line[idx[col]] = str(rec["counts"][col])
        for col in WEIGHT_AUTHORED:
            line[idx[col]] = _fmt_avg(rec["file_w"][col], rec["file_n"][col])
        line[idx["min_hours_pr_created_to_merged"]] = (
            f"{rec['min_hours']['min_hours_pr_created_to_merged']:.4f}"
            if "min_hours_pr_created_to_merged" in rec["min_hours"]
            else ""
        )
        line[idx["max_hours_pr_created_to_merged"]] = (
            f"{rec['max_hours']['max_hours_pr_created_to_merged']:.4f}"
            if "max_hours_pr_created_to_merged" in rec["max_hours"]
            else ""
        )
        line[idx["avg_hours_pr_created_to_merged"]] = _fmt_avg(
            rec["merge_w"]["avg_hours_pr_created_to_merged"],
            rec["merge_n"]["avg_hours_pr_created_to_merged"],
        )
        out.append(line)

    return headers, out


def _empty_totals_row(headers: list[str], user: str) -> list[str]:
    idx = {h: i for i, h in enumerate(headers)}
    row = [""] * len(headers)
    row[idx["user"]] = user
    for col in COUNT_COLUMNS:
        if col in idx:
            row[idx[col]] = "0"
    return row


def _team_sheet_from_totals(
    totals_headers: list[str],
    totals_rows: list[list[str]],
    team: Team,
) -> tuple[list[str], list[list[str]]]:
    """Filter Totals rows for team members; include roster order and staff_name."""
    user_idx = totals_headers.index("user")
    by_login = {row[user_idx].strip().lower(): row for row in totals_rows}
    headers = ["staff_name", *totals_headers]
    rows: list[list[str]] = []
    for member in team.members:
        src = by_login.get(member.user.strip().lower())
        if src is None:
            src = _empty_totals_row(totals_headers, member.user)
        rows.append([member.name, *src])
    return headers, rows


def _cell_value(raw: str):
    """Prefer real numbers in Excel (avoids 'number stored as text' warnings)."""
    s = (raw or "").strip()
    if not s:
        return ""
    try:
        if re.fullmatch(r"-?\d+", s):
            return int(s)
        if re.fullmatch(r"-?\d+\.\d+", s):
            return float(s)
    except ValueError:
        pass
    return s


def _write_sheet(
    workbook,
    title: str,
    headers: list[str],
    rows: list[list[str]],
    *,
    active: bool = False,
) -> None:
    from openpyxl.styles import Font

    if active:
        ws = workbook.active
        ws.title = _sheet_title(title)
    else:
        ws = workbook.create_sheet(_sheet_title(title))

    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    _autosize_columns(ws, headers, rows)


TOTALS_BANNER = "All Staff Metrics"


def _write_bannered_sheet(
    workbook,
    title: str,
    banner: str,
    headers: list[str],
    rows: list[list[str]],
    *,
    active: bool = False,
) -> None:
    """Bannered sheet: title banner, underlined headers, roomier cells, numeric values."""
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    if active:
        ws = workbook.active
        ws.title = _sheet_title(title)
    else:
        ws = workbook.create_sheet(_sheet_title(title))

    last_col = max(len(headers), 1)
    header_row = 2
    data_start = 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    banner_cell = ws.cell(row=1, column=1, value=banner)
    banner_cell.font = Font(bold=True, size=14)
    banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    underline = Border(bottom=Side(style="medium", color="000000"))
    col_align = Alignment(horizontal="left", vertical="center", indent=1)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = underline
        cell.alignment = col_align
    ws.row_dimensions[header_row].height = 22

    for r_i, row in enumerate(rows, start=data_start):
        ws.row_dimensions[r_i].height = 20
        for c_i, value in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=_cell_value(value))
            cell.alignment = col_align

    widths = [len(h) for h in headers]
    for row in rows:
        for idx, value in enumerate(row):
            if idx < len(widths):
                widths[idx] = max(widths[idx], len(str(value)))
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 4, 12), 64)


def combine_workbook(
    summary_paths: list[Path],
    output_path: Path,
    *,
    repo_order: list[str] | None = None,
    teams: list[Team] | None = None,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit("openpyxl required: uv sync --group excel") from exc

    if not summary_paths:
        raise SystemExit("no *_person_summary.tsv files found")

    labeled: list[tuple[str, Path]] = []
    for path in summary_paths:
        label = _repo_label_from_summary_path(path)
        labeled.append((label, path))

    if repo_order:
        order_map = {name: i for i, name in enumerate(repo_order)}
        labeled.sort(key=lambda t: (order_map.get(t[0], 999), t[0].lower()))
    else:
        labeled.sort(key=lambda t: t[0].lower())

    repo_rows: list[tuple[str, list[str], list[list[str]]]] = []
    headers: list[str] | None = None
    for label, path in labeled:
        h, rows = read_tsv_table(path)
        if headers is None:
            headers = h
        elif h != headers:
            raise SystemExit(f"header mismatch in {path}")
        repo_rows.append((label, h, rows))

    assert headers is not None
    totals_headers, totals_rows = _aggregate_totals(
        headers, [(label, rows) for label, _, rows in repo_rows]
    )

    workbook = Workbook()
    _write_bannered_sheet(
        workbook,
        "Totals",
        TOTALS_BANNER,
        totals_headers,
        totals_rows,
        active=True,
    )

    for team in teams or []:
        th, tr = _team_sheet_from_totals(totals_headers, totals_rows, team)
        _write_bannered_sheet(
            workbook,
            team.name,
            team.team_banner or team.name,
            th,
            tr,
        )

    for label, h, rows in repo_rows:
        _write_sheet(workbook, label, h, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    from github_analysis.config import OUTPUT_DIR_HELP, default_output_dir, resolve_output_dir

    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Output directory precedence: --output-dir → GITHUB_ANALYSIS_RESULTS → "
            f"built-in default ({default_output_dir()}).\n"
            "CLI --output-dir overrides the environment variable when both are set.\n"
            f"Team tabs: --teams-config (default {DEFAULT_TEAMS_CONFIG} if present)."
        ),
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        help="Directory containing *_person_summary.tsv files (default: --output-dir)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help=OUTPUT_DIR_HELP,
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Combined Excel output path (.xlsx). Required unless --output-dir and --combined-name are set",
    )
    parser.add_argument(
        "--combined-name",
        default="",
        help="Filename under --output-dir when -o is omitted (e.g. combined-2026-07.xlsx)",
    )
    parser.add_argument(
        "--repo-order",
        nargs="*",
        default=[
            "global-services",
            "global-user-services",
            "polaris-turbo",
            "org-manager",
            "places-proxy",
        ],
        help="Sheet order after Totals/team tabs (default: CEDT repos)",
    )
    parser.add_argument(
        "--teams-config",
        type=Path,
        default=None,
        help=(
            "YAML or JSON team roster for agency tabs "
            f"(default: {DEFAULT_TEAMS_CONFIG} if it exists; use --no-teams to skip)"
        ),
    )
    parser.add_argument(
        "--no-teams",
        action="store_true",
        help="Do not add team tabs even if a teams config exists",
    )
    parser.add_argument(
        "--stem-suffix",
        default="",
        help=(
            "Only include summaries whose filename contains this string "
            "(e.g. 2026-06-01_to_2026-06-23). Required when --input-dir holds "
            "multiple date windows."
        ),
    )
    args = parser.parse_args()

    out_dir = Path(resolve_output_dir(str(args.output_dir) if args.output_dir else None)).expanduser()
    input_dir = (args.input_dir or out_dir).expanduser()

    if args.output:
        output_path = args.output.expanduser()
    elif args.combined_name:
        output_path = out_dir / args.combined_name
    else:
        raise SystemExit("error: specify -o/--output or --output-dir with --combined-name")

    paths = sorted(input_dir.glob("*_person_summary.tsv"))
    if args.stem_suffix:
        paths = [p for p in paths if args.stem_suffix in p.name]
        if not paths:
            raise SystemExit(
                f"no *_person_summary.tsv matching stem-suffix {args.stem_suffix!r} in {input_dir}"
            )

    teams: list[Team] = []
    if not args.no_teams:
        cfg = args.teams_config
        if cfg is None and DEFAULT_TEAMS_CONFIG.is_file():
            cfg = DEFAULT_TEAMS_CONFIG
        if cfg is not None:
            teams = load_teams_config(cfg)
            print(f"Team tabs: {', '.join(t.name for t in teams)} (from {cfg})")

    combine_workbook(paths, output_path, repo_order=args.repo_order, teams=teams)
    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
